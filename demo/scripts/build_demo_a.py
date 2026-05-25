"""Demo A: 顧客A_引継ぎ資料.xlsx 生成スクリプト

シナリオ：建設会社「ピーエス・コンストラクション」の営業担当・田中が異動。
後任の佐藤が、架空の顧客A（東洋ロジスティクス株式会社）との過去の交渉を
引き継ぐためのデータを Claude for Work に添付して使うことを想定。

実行: python3 demo/scripts/build_demo_a.py
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

OUTPUT = Path(__file__).resolve().parent.parent / "demo-a-handover" / "顧客A_引継ぎ資料.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="0A3D62")
HEADER_FONT = Font(name="游ゴシック", size=11, bold=True, color="FFFFFF")
BODY_FONT = Font(name="游ゴシック", size=10)
BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        cell.border = BORDER


def style_body(ws, start_row, end_row, cols):
    for r in range(start_row, end_row + 1):
        for c in range(1, cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = BORDER


def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_case_overview(wb):
    ws = wb.create_sheet("案件概要")
    rows = [
        ("項目", "内容"),
        ("顧客名", "東洋ロジスティクス株式会社"),
        ("顧客所在地", "東京都江東区青海2-X-X"),
        ("顧客業種", "物流（3PL・倉庫運営）"),
        ("顧客規模", "従業員 約1,200名 / 売上 約450億円"),
        ("案件名", "千葉県市原市 物流倉庫 新築工事"),
        ("案件規模", "延床面積 約32,000㎡ / 鉄骨造一部PCa / 3階建"),
        ("案件フェーズ", "実施設計フェーズ・見積条件再調整中"),
        ("予算規模", "概算 48〜52億円（最終調整中）"),
        ("竣工希望", "2026年12月（先方希望、2027年3月までは許容）"),
        ("PSC担当営業（旧）", "田中 健司（2026年5月に名古屋支店へ異動）"),
        ("PSC担当営業（新）", "佐藤 真理子（本資料の引継ぎ対象）"),
        ("PSC担当設計", "山本 慎二（継続）"),
        ("先方主担当", "営業企画部 部長 / 加納 利明 様"),
        ("先方決裁者", "取締役 経営戦略本部長 / 西郷 英樹 様"),
        ("商談ステータス", "競合2社と並走中（C社・D社）。技術提案で優位、価格で劣勢"),
        ("次回打ち合わせ", "2026年6月2日(火) 14:00 〜 先方本社"),
    ]
    for r, (k, v) in enumerate(rows, start=1):
        ws.cell(row=r, column=1, value=k)
        ws.cell(row=r, column=2, value=v)
    style_header(ws, 1, 2)
    style_body(ws, 2, len(rows), 2)
    set_widths(ws, [22, 70])
    ws.freeze_panes = "A2"


def build_emails(wb):
    ws = wb.create_sheet("メール履歴")
    header = ["No", "日付", "送信者", "宛先", "件名", "本文"]
    ws.append(header)
    emails = [
        (1, "2024-04-15", "加納 利明（東洋ロジ）", "田中 健司（PSC）",
         "物流倉庫新築のお見積りご相談",
         "千葉県市原市での物流倉庫新築を計画しています。延床32,000㎡規模・3階建・一部PCa構造を想定。来週どこかでお打合せいただけませんか。"),
        (2, "2024-05-10", "田中 健司（PSC）", "加納 利明（東洋ロジ）",
         "ご相談内容確認のうえ初回提案資料",
         "ご相談ありがとうございます。PCa構造の優位性と弊社実績資料をお送りします。来週5/20午後でいかがでしょうか。"),
        (3, "2024-06-03", "加納 利明（東洋ロジ）", "田中 健司（PSC）",
         "屋根荷重条件について",
         "屋上に太陽光パネル（約180kW）を載せたい要望が出ました。屋根の積載荷重を上げる方向で見積条件を見直してください。"),
        (4, "2024-07-22", "加納 利明（東洋ロジ）", "田中 健司（PSC）",
         "仕様変更のご相談（断熱・冷蔵区画追加）",
         "1階の一部（約3,000㎡）を冷蔵倉庫区画に変更したいです。断熱仕様と床荷重の見直しをお願いします。"),
        (5, "2024-08-01", "田中 健司（PSC）", "加納 利明（東洋ロジ）",
         "冷蔵区画追加についての概算影響",
         "冷蔵区画追加で概算+3.2億円のインパクトです。詳細は8/15会議でご説明します。"),
        (6, "2024-09-03", "加納 利明（東洋ロジ）", "田中 健司（PSC）",
         "予算上限のご共有と価格相談",
         "経営会議の結果、予算上限は52億円となりました。現在の提示額（54.8億円）を3億円圧縮できないか相談です。"),
        (7, "2024-09-20", "田中 健司（PSC）", "加納 利明（東洋ロジ）",
         "コスト圧縮案",
         "外壁仕様の一部変更・空調系の機器グレード見直しで2.4億円の圧縮案を作成しました。冷蔵区画の仕様はキープです。"),
        (8, "2024-11-10", "加納 利明（東洋ロジ）", "田中 健司（PSC）",
         "競合C社提案の話",
         "C社から49億円の提案が来ています。価格では負けていますが、御社の工期と品質に期待しています。技術提案を強化してほしいです。"),
        (9, "2025-01-15", "西郷 英樹（東洋ロジ）", "田中 健司（PSC）",
         "経営層への直接ご説明依頼",
         "次回、私（西郷）も同席します。竣工後の運用コスト・耐震性能・PCa工法の優位性を経営層向けに整理してください。"),
        (10, "2025-03-05", "田中 健司（PSC）", "加納 利明 / 西郷 英樹（東洋ロジ）",
         "技術提案書 v3 ご送付",
         "経営層向け資料を含む技術提案書v3をお送りします。運用コスト（30年）・耐震性能（BCP指標）・近隣対応プランを追加しました。"),
        (11, "2025-04-22", "加納 利明（東洋ロジ）", "田中 健司（PSC）",
         "近隣対応について",
         "計画地の北側住宅街への騒音・振動対策で住民から懸念が出ています。近隣説明会のサポートをお願いできますか。"),
        (12, "2025-05-10", "加納 利明（東洋ロジ）", "田中 健司（PSC） / 佐藤 真理子（PSC）",
         "担当変更のご連絡を受けて",
         "田中さんの異動の件、承知しました。佐藤様、6/2の打ち合わせでよろしくお願いします。事前に資料一式お送りいただけると助かります。"),
    ]
    for em in emails:
        ws.append(em)
    style_header(ws, 1, len(header))
    style_body(ws, 2, len(emails) + 1, len(header))
    set_widths(ws, [5, 12, 22, 22, 36, 70])
    ws.freeze_panes = "A2"


def build_meetings(wb):
    ws = wb.create_sheet("議事録")
    header = ["No", "日付", "場所", "参加者", "議題", "決定事項", "宿題・未決"]
    ws.append(header)
    meetings = [
        (1, "2024-05-20", "東洋ロジ本社", "加納（東洋ロジ）／田中・山本（PSC）",
         "初回打合せ。要件確認と工法説明",
         "PCa工法を主案で進める。屋根は太陽光パネル想定なし（時点）",
         "5月末までに概算予算とスケジュール提示（PSC）"),
        (2, "2024-08-15", "PSC本社",
         "加納（東洋ロジ）／田中・山本・空調担当（PSC）",
         "仕様変更（冷蔵区画追加）の確認",
         "冷蔵区画3,000㎡を1階北側に配置。断熱仕様A種。床荷重1.5t/㎡",
         "冷蔵区画追加の概算+3.2億円の根拠資料を9/10までに提出（PSC）。先方経営会議で予算上限を確認"),
        (3, "2024-11-25", "東洋ロジ本社",
         "加納・経理部長（東洋ロジ）／田中・山本（PSC）",
         "コスト圧縮案レビュー",
         "外壁の一部仕様変更を採用（▲1.4億円）。空調機器グレード見直しを採用（▲1.0億円）",
         "圧縮後の最終見積（52.4億円）の妥当性検証（先方）。BCP耐震性能の追加資料（PSC）"),
        (4, "2025-03-12", "東洋ロジ本社",
         "西郷取締役・加納（東洋ロジ）／田中・山本（PSC）",
         "経営層向け技術提案レビュー",
         "30年運用コスト試算とBCP耐震性能の説明に好印象。PCa工法の工期短縮メリットを評価",
         "近隣（北側住宅）対策案の具体化（PSC）。価格面での再見直しは行わない方針（先方）"),
    ]
    for m in meetings:
        ws.append(m)
    style_header(ws, 1, len(header))
    style_body(ws, 2, len(meetings) + 1, len(header))
    set_widths(ws, [5, 12, 18, 36, 28, 40, 40])
    ws.freeze_panes = "A2"


def build_requirements(wb):
    ws = wb.create_sheet("要望リスト")
    header = ["No", "要望項目", "初出日", "最新ステータス", "優先度", "備考"]
    ws.append(header)
    reqs = [
        (1, "屋上太陽光パネル設置（180kW）", "2024-06-03", "見積条件に反映済（屋根積載荷重UP）", "高",
         "経営層のESG施策と連動。減らせない要望"),
        (2, "1F北側に冷蔵倉庫区画（3,000㎡）追加", "2024-07-22", "仕様確定済", "高",
         "断熱A種・床荷重1.5t/㎡・庫内温度+5℃想定"),
        (3, "総予算 52億円以内", "2024-09-03", "圧縮案で52.4億円まで圧縮、最終調整中", "高",
         "経営会議で決定済の上限"),
        (4, "竣工 2026年12月希望", "2024-04-15", "実現可能（PCa工法の工期短縮で）", "中",
         "2027年3月までは許容"),
        (5, "BCP耐震性能の証拠資料", "2025-01-15", "技術提案書v3に追加済", "高",
         "経営層が重視。継続的にアップデート要"),
        (6, "北側住宅街への騒音・振動対策", "2025-04-22", "近隣対応プラン作成中", "高",
         "**未解決**。説明会の段取り含めPSCにサポート期待"),
        (7, "30年運用コスト試算の継続提供", "2025-03-05", "経営層に好評", "中",
         "竣工後も含めた長期視点での提案を期待"),
        (8, "C社（競合）対比での技術優位の継続説明", "2024-11-10", "技術提案書v3で対応", "中",
         "価格では負けているため、技術で勝つ必要"),
    ]
    for r in reqs:
        ws.append(r)
    style_header(ws, 1, len(header))
    style_body(ws, 2, len(reqs) + 1, len(header))
    set_widths(ws, [5, 30, 12, 32, 8, 38])
    ws.freeze_panes = "A2"


def build_keypersons(wb):
    ws = wb.create_sheet("キーパーソン")
    header = ["氏名", "所属・役職", "関係性", "コミュニケーション特性", "注意点・メモ"]
    ws.append(header)
    persons = [
        ("加納 利明 様", "東洋ロジ 営業企画部 部長",
         "主担当窓口。日常的なやり取り",
         "メール多用・即レス派・数字に強い",
         "急ぎの相談は電話の方が好まれる。技術論より経営インパクトを好む"),
        ("西郷 英樹 様", "東洋ロジ 取締役 経営戦略本部長",
         "決裁者。要所で同席",
         "対面重視・抽象論を嫌う・ROI/長期視点を重視",
         "資料は1枚で要点。BCP・運用コスト・ESGに反応する"),
        ("石川 弘樹 様", "東洋ロジ 経理部長",
         "コスト面の最終チェック",
         "詳細詰め型・数字の根拠を必ず求める",
         "見積根拠の出典を必ず添える"),
        ("田中 健司", "PSC 営業（前任）",
         "前任。引継ぎ元",
         "—", "現在は名古屋支店。緊急時は電話で連絡可"),
        ("佐藤 真理子", "PSC 営業（後任）",
         "本資料の引継ぎ対象",
         "—", "—"),
        ("山本 慎二", "PSC 設計",
         "継続担当。技術提案のキーパーソン",
         "—", "顧客との信頼関係構築済み。佐藤同行が安心"),
    ]
    for p in persons:
        ws.append(p)
    style_header(ws, 1, len(header))
    style_body(ws, 2, len(persons) + 1, len(header))
    set_widths(ws, [16, 32, 28, 32, 42])
    ws.freeze_panes = "A2"


def main():
    wb = Workbook()
    wb.remove(wb.active)
    build_case_overview(wb)
    build_emails(wb)
    build_meetings(wb)
    build_requirements(wb)
    build_keypersons(wb)
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    print(f"wrote: {OUTPUT}")


if __name__ == "__main__":
    main()

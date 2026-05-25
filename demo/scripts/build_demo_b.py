"""Demo B: 社内図説シート_テンプレート.xlsx 生成スクリプト

仕様書PDFから読み込んだ内容を流し込む空テンプレート。
凡例・記入要領・図説シートの3シート構成。

実行: python3 demo/scripts/build_demo_b.py
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

OUTPUT = Path(__file__).resolve().parent.parent / "demo-b-docgen" / "社内図説シート_テンプレート.xlsx"

NAVY = "0A3D62"
SOFT = "F5F7FA"
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
SOFT_FILL = PatternFill("solid", fgColor=SOFT)
HEADER_FONT = Font(name="游ゴシック", size=11, bold=True, color="FFFFFF")
SUB_FONT = Font(name="游ゴシック", size=10, bold=True)
BODY_FONT = Font(name="游ゴシック", size=10)
BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def style_header_row(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        cell.border = BORDER


def style_body(ws, start_row, end_row, cols, fill=None):
    for r in range(start_row, end_row + 1):
        for c in range(1, cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = BORDER
            if fill is not None:
                cell.fill = fill


def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_zusetsu(wb):
    ws = wb.create_sheet("図説シート")
    header = ["No", "工種コード", "工種", "項目", "仕様", "数量", "単位", "備考", "出典（仕様書ページ・見出し）"]
    ws.append(header)
    style_header_row(ws, 1, len(header))
    # 例示の1行だけ薄グレーで残す（記入例）
    example = [1, "C-01", "コンクリート工事", "基礎コンクリート",
               "（仕様書から記入）", "（数量）", "（単位）",
               "（特記事項）", "（出典：例 P.12 §3.2）"]
    ws.append(example)
    style_body(ws, 2, 2, len(header), fill=SOFT_FILL)
    # 残り20行は空のまま罫線だけ
    for _ in range(20):
        ws.append([""] * len(header))
    style_body(ws, 3, 22, len(header))
    set_widths(ws, [5, 12, 18, 22, 32, 10, 8, 28, 26])
    ws.freeze_panes = "A2"


def build_rules(wb):
    ws = wb.create_sheet("記入要領")
    rows = [
        ("項目", "記入ルール"),
        ("工種コード", "凡例シートのコード体系に従う。例：C-XX（コンクリート系）、S-XX（鉄骨系）、F-XX（仕上系）"),
        ("工種", "凡例シートの工種名から選択"),
        ("項目", "仕様書の見出しに合わせる。複数項目は行を分ける"),
        ("仕様", "材質・グレード・寸法・性能を明記。例：「FC=30N/mm² スランプ18cm」"),
        ("数量", "数値のみ。単位は別列。小数点以下は2桁まで"),
        ("単位", "凡例シートの単位略号から選択（㎡・㎥・本・kg 等）"),
        ("備考", "特記事項・先方確認事項・社内補足を記入"),
        ("出典", "仕様書のページ番号と見出しを必ず明記。AIたたき台でも引用ベースで記載"),
        ("不明な値の扱い", "推測しない。「要確認」と記入し、出典に「不明」と明記"),
    ]
    for k, v in rows:
        ws.append([k, v])
    style_header_row(ws, 1, 2)
    style_body(ws, 2, len(rows), 2)
    set_widths(ws, [20, 80])
    ws.freeze_panes = "A2"


def build_legend(wb):
    ws = wb.create_sheet("凡例")
    # 工種コード
    ws.cell(row=1, column=1, value="工種コード一覧").font = SUB_FONT
    ws.append(["コード", "工種名"])
    coverage = [
        ("C-01", "コンクリート工事（基礎）"),
        ("C-02", "コンクリート工事（躯体）"),
        ("C-03", "プレキャストコンクリート（PCa）部材"),
        ("C-04", "プレストレストコンクリート（PC）工事"),
        ("S-01", "鉄骨工事（柱・梁）"),
        ("S-02", "鉄筋工事"),
        ("F-01", "仕上工事（外装）"),
        ("F-02", "仕上工事（内装）"),
        ("M-01", "設備工事（電気）"),
        ("M-02", "設備工事（空調）"),
        ("M-03", "設備工事（給排水）"),
        ("X-01", "外構工事"),
    ]
    for code, name in coverage:
        ws.append([code, name])
    style_header_row(ws, 2, 2)
    style_body(ws, 3, 2 + len(coverage), 2)

    # 単位略号
    start = 2 + len(coverage) + 2
    ws.cell(row=start, column=1, value="単位略号一覧").font = SUB_FONT
    ws.cell(row=start + 1, column=1, value="単位")
    ws.cell(row=start + 1, column=2, value="意味")
    style_header_row(ws, start + 1, 2)
    units = [
        ("㎡", "平方メートル"),
        ("㎥", "立方メートル"),
        ("m", "メートル"),
        ("本", "本"),
        ("基", "基"),
        ("kg", "キログラム"),
        ("t", "トン"),
        ("式", "一式"),
    ]
    for u, m in units:
        ws.append([u, m])
    style_body(ws, start + 2, start + 1 + len(units), 2)

    set_widths(ws, [14, 40])


def main():
    wb = Workbook()
    wb.remove(wb.active)
    build_zusetsu(wb)
    build_rules(wb)
    build_legend(wb)
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    print(f"wrote: {OUTPUT}")


if __name__ == "__main__":
    main()

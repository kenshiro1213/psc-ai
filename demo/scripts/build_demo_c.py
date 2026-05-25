"""Demo C: 発注公告_収集データ.xlsx 生成スクリプト

複数自治体から手作業で収集したと仮定した発注公告データ。
シナリオ：このExcelをClaude for Workに添付して、PSCの関心条件で絞り込み・整形させる。

実行: python3 demo/scripts/build_demo_c.py
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

OUTPUT = Path(__file__).resolve().parent.parent / "demo-c-bid-watch" / "発注公告_収集データ.xlsx"

NAVY = "0A3D62"
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
HEADER_FONT = Font(name="游ゴシック", size=11, bold=True, color="FFFFFF")
BODY_FONT = Font(name="游ゴシック", size=10)
BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        cell.border = BORDER


def style_body(ws, start_row, end_row, cols):
    for r in range(start_row, end_row + 1):
        for c in range(1, cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = BORDER


def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_announcements(wb):
    ws = wb.create_sheet("公告データ")
    header = ["No", "自治体", "公告日", "案件名", "工種", "予定価格(百万円)", "場所", "工期(月)", "URL", "公告本文（抜粋）"]
    ws.append(header)
    rows = [
        (1, "千葉県市原市", "2026-05-12", "市原市立第三中学校 体育館 改築工事", "建築（PCa可）", 820, "千葉県市原市", 14, "https://example.go.jp/ichihara/01", "鉄骨造一部PCa構造、延床1,800㎡。耐震性能・工期短縮の提案を歓迎。"),
        (2, "千葉県市原市", "2026-05-14", "上下水道 配管更新工事 第3期", "土木（配管）", 240, "千葉県市原市", 8, "https://example.go.jp/ichihara/02", "既設管路の更新。狭隘部の施工計画が課題。"),
        (3, "東京都江東区", "2026-05-15", "区立体育館 屋根改修工事", "建築（改修）", 180, "東京都江東区", 6, "https://example.go.jp/koto/01", "屋根の防水・断熱改修。営業しながらの工事。"),
        (4, "東京都江東区", "2026-05-16", "区民会館 新築工事（PCa工法可）", "建築（PCa可）", 1850, "東京都江東区", 22, "https://example.go.jp/koto/02", "RC造一部PCa。工期短縮の提案歓迎。BCP耐震性能を重視。"),
        (5, "神奈川県横浜市", "2026-05-13", "市営住宅 耐震補強工事", "建築（耐震）", 520, "神奈川県横浜市", 10, "https://example.go.jp/yokohama/01", "既存RC造住宅の耐震補強。住民居住中。"),
        (6, "神奈川県横浜市", "2026-05-17", "新港物流ターミナル PCa床版製作", "PCa製作", 380, "神奈川県横浜市", 7, "https://example.go.jp/yokohama/02", "桟橋上部工のPCa床版製作・据付。"),
        (7, "埼玉県さいたま市", "2026-05-12", "市立図書館 新築工事", "建築", 950, "埼玉県さいたま市", 18, "https://example.go.jp/saitama/01", "鉄骨造3階建。延床2,400㎡。"),
        (8, "埼玉県さいたま市", "2026-05-18", "下水管 更生工事 第5期", "土木（更生）", 150, "埼玉県さいたま市", 5, "https://example.go.jp/saitama/02", "既設管の更生（非開削）。延長約1.8km。"),
        (9, "茨城県つくば市", "2026-05-11", "つくばエクスプレス沿線 物流倉庫 公募", "建築（倉庫）", 4200, "茨城県つくば市", 24, "https://example.go.jp/tsukuba/01", "延床32,000㎡規模。鉄骨造一部PCa。冷蔵区画含む。"),
        (10, "茨城県水戸市", "2026-05-13", "市道アンダーパス 耐震改修", "土木（耐震）", 310, "茨城県水戸市", 9, "https://example.go.jp/mito/01", "既設アンダーパスの耐震性能向上。"),
        (11, "栃木県宇都宮市", "2026-05-15", "市民プール 改築工事", "建築（改築）", 680, "栃木県宇都宮市", 16, "https://example.go.jp/utsunomiya/01", "プール屋根の大スパン構造。"),
        (12, "群馬県前橋市", "2026-05-16", "市立小学校 トイレ改修", "建築（改修）", 75, "群馬県前橋市", 4, "https://example.go.jp/maebashi/01", "小規模改修。"),
        (13, "新潟県新潟市", "2026-05-12", "港湾 PCa桟橋 製作据付", "PCa製作", 1180, "新潟県新潟市", 18, "https://example.go.jp/niigata/01", "プレキャストプレストレスト桟橋。"),
        (14, "静岡県静岡市", "2026-05-14", "県立病院 駐車場棟 新築", "建築（駐車場）", 720, "静岡県静岡市", 13, "https://example.go.jp/shizuoka/01", "鉄骨造一部PCa。BCP対応。"),
        (15, "愛知県名古屋市", "2026-05-13", "市営住宅 集約建替 第2期", "建築（集合住宅）", 1620, "愛知県名古屋市", 20, "https://example.go.jp/nagoya/01", "RC造一部PCa。延床18,000㎡。"),
        (16, "愛知県名古屋市", "2026-05-17", "下水処理場 増設工事", "土木（プラント）", 2400, "愛知県名古屋市", 28, "https://example.go.jp/nagoya/02", "既設稼働中の処理場の増設。"),
        (17, "三重県四日市市", "2026-05-15", "市役所 耐震補強・設備更新", "建築（耐震）", 540, "三重県四日市市", 12, "https://example.go.jp/yokkaichi/01", "既存庁舎の耐震・設備一括更新。"),
        (18, "大阪府大阪市", "2026-05-12", "舞洲物流センター 新築（PCa）", "建築（倉庫）", 5800, "大阪府大阪市", 26, "https://example.go.jp/osaka/01", "延床45,000㎡。鉄骨造一部PCa。"),
        (19, "京都府京都市", "2026-05-14", "歴史的建造物 補強工事", "建築（特殊）", 410, "京都府京都市", 14, "https://example.go.jp/kyoto/01", "文化財指定建造物の構造補強。"),
        (20, "兵庫県神戸市", "2026-05-16", "ポートライナー高架 補修", "土木（高架）", 980, "兵庫県神戸市", 16, "https://example.go.jp/kobe/01", "既設高架橋の補修・補強。"),
        (21, "広島県広島市", "2026-05-13", "市立スポーツセンター 改修", "建築（改修）", 260, "広島県広島市", 7, "https://example.go.jp/hiroshima/01", "既存施設の機能更新。"),
        (22, "福岡県福岡市", "2026-05-15", "市営住宅 PCa工法 集約建替", "建築（集合住宅）", 1480, "福岡県福岡市", 19, "https://example.go.jp/fukuoka/01", "RC造一部PCa。延床16,000㎡。工期短縮重視。"),
        (23, "福岡県北九州市", "2026-05-17", "若松物流倉庫 新築", "建築（倉庫）", 2650, "福岡県北九州市", 22, "https://example.go.jp/kitakyushu/01", "延床28,000㎡。鉄骨造一部PCa。"),
        (24, "熊本県熊本市", "2026-05-14", "震災復興住宅 PCa工法", "建築（集合住宅）", 760, "熊本県熊本市", 14, "https://example.go.jp/kumamoto/01", "PCa工法による工期短縮重視。"),
        (25, "鹿児島県鹿児島市", "2026-05-16", "桜島フェリーターミナル 改修", "建築（改修）", 320, "鹿児島県鹿児島市", 8, "https://example.go.jp/kagoshima/01", "営業中の改修。"),
        (26, "北海道札幌市", "2026-05-12", "地下鉄駅 リフレッシュ工事", "土木（駅施設）", 480, "北海道札幌市", 11, "https://example.go.jp/sapporo/01", "既存駅構内の改修。夜間施工含む。"),
        (27, "北海道千歳市", "2026-05-15", "千歳基地 PCa格納庫 増設", "建築（特殊）", 1240, "北海道千歳市", 18, "https://example.go.jp/chitose/01", "PCa工法・大スパン構造。"),
        (28, "宮城県仙台市", "2026-05-13", "仙台港 物流倉庫 新築", "建築（倉庫）", 3200, "宮城県仙台市", 24, "https://example.go.jp/sendai/01", "延床36,000㎡。鉄骨造一部PCa。BCP対応。"),
        (29, "宮城県仙台市", "2026-05-17", "市道橋梁 PCa桁 製作", "PCa製作", 580, "宮城県仙台市", 12, "https://example.go.jp/sendai/02", "プレキャストPC桁の製作。"),
        (30, "福島県郡山市", "2026-05-16", "市立体育館 屋根改修", "建築（改修）", 140, "福島県郡山市", 5, "https://example.go.jp/koriyama/01", "屋根の防水改修。"),
    ]
    for r in rows:
        ws.append(r)
    style_header(ws, 1, len(header))
    style_body(ws, 2, len(rows) + 1, len(header))
    set_widths(ws, [5, 18, 12, 38, 16, 14, 16, 8, 32, 60])
    ws.freeze_panes = "A2"


def build_criteria(wb):
    ws = wb.create_sheet("抽出条件")
    rows = [
        ("項目", "条件", "備考"),
        ("工種キーワード", "PCa, プレキャスト, プレストレスト, 物流倉庫, BCP",
         "PSCの得意領域。OR条件で1つでも含めば候補"),
        ("予定価格レンジ", "500百万円 〜 6,000百万円",
         "下限以下は採算性、上限超は他社と組む案件"),
        ("地域", "関東圏（東京・千葉・神奈川・埼玉・茨城・栃木・群馬）優先、その他は要相談",
         "営業の主軸が関東"),
        ("除外キーワード", "改修、補強、補修、トイレ、屋根改修のみ、小規模",
         "新築・本体工事を優先"),
        ("工期下限", "10ヶ月以上",
         "短工期案件は工程が厳しい"),
        ("特に注目", "BCP、耐震性能、太陽光、PCa工法、工期短縮を要件に含む案件",
         "PSCの強みが活きる"),
    ]
    for r in rows:
        ws.append(r)
    style_header(ws, 1, 3)
    style_body(ws, 2, len(rows), 3)
    set_widths(ws, [22, 56, 40])
    ws.freeze_panes = "A2"


def build_output_template(wb):
    ws = wb.create_sheet("出力テンプレ")
    # シンプルなテキストブロック
    ws.cell(row=1, column=1, value="共有メール／Slack文面テンプレート").font = Font(name="游ゴシック", size=12, bold=True)
    template = """件名：【発注公告ウォッチ】今週の関心案件 {{件数}}件（{{集計日}}）

お疲れさまです。
今週の自治体発注公告から、PSCの関心条件に合致する案件を抽出しました。

【ハイライト】
{{ハイライト1行コメント}}

【関心案件一覧】
{{#each 案件}}
■ {{案件名}}（{{自治体}}）
  ・工種：{{工種}}
  ・予定価格：{{予定価格}}百万円
  ・工期：{{工期}}ヶ月
  ・場所：{{場所}}
  ・PSC適合理由：{{適合理由}}
  ・URL：{{URL}}
{{/each}}

【次のアクション】
{{推奨アクション}}

──
集計：自動巡回スクリプト（PoC運用中）"""
    ws.cell(row=2, column=1, value=template)
    ws.cell(row=2, column=1).alignment = Alignment(vertical="top", wrap_text=True)
    ws.cell(row=2, column=1).font = BODY_FONT
    ws.row_dimensions[2].height = 360
    ws.column_dimensions["A"].width = 96


def main():
    wb = Workbook()
    wb.remove(wb.active)
    build_announcements(wb)
    build_criteria(wb)
    build_output_template(wb)
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    print(f"wrote: {OUTPUT}")


if __name__ == "__main__":
    main()
